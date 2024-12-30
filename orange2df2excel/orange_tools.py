from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.dataframe import dataframe_to_rows
import xlsxwriter
from koboextractor import KoboExtractor
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives import padding
from cryptography.hazmat.backends import default_backend
from Crypto.Random import get_random_bytes
from Crypto.Protocol.KDF import PBKDF2
from Crypto.Cipher import AES
from Crypto.Util.Padding import pad, unpad
from Crypto.Protocol.KDF import PBKDF2
import base64
import requests
import pandas as pd
import os
from io import StringIO
import hashlib
import bcrypt

def raw_data_to_excel(df, file_path, sheet_name):
    """
    Write a DataFrame to an Excel file in table format.
    
    Parameters:
    - df: pandas.DataFrame - The DataFrame to write to Excel.
    - file_path: str - Path to the Excel file.
    - sheet_name: str - Name of the sheet to write data to.
    """
    if os.path.exists(file_path):
        workbook = load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]
    else:
        workbook = Workbook()
        if 'Sheet' in workbook.sheetnames:
            del workbook['Sheet']

    worksheet = workbook.create_sheet(sheet_name)

    for row in dataframe_to_rows(df, index=False, header=True):
        worksheet.append(row)

    table = Table(displayName="raw_data", ref=worksheet.dimensions)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True
    )
    table.tableStyleInfo = style
    worksheet.add_table(table)

#-----------Adjusting cells--------------
    for col in worksheet.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Get the column letter
        for cell in col:
            if cell.value:
                # Estimate width by multiplying character count by a width factor
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[col_letter].width = adjusted_width

    workbook.save(file_path)

def raw_data_to_excel_with_all_charts(df, file_path, chart_config, totals=None):
    """
    Write raw data to an Excel file and create a clean dashboard with various chart types using `xlsxwriter`.

    Parameters:
    - df: pandas.DataFrame - The data to write to the Excel file.
    - file_path: str - Path to save the Excel file.
    - chart_config: dict - Dictionary to configure charts.
        Keys are chart types (e.g., "bar", "line", "pie").
        Values are dicts with keys:
            - 'category_col': str - Column to use as categories.
            - 'value_col': str - Column to use as values.
    - totals: list - List of column names to calculate totals for. If None, totals will not be shown.
    """
    with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
        df.to_excel(writer, sheet_name='Raw Data', index=False)
        workbook = writer.book
        dashboard = workbook.add_worksheet('Dashboard')
        row_offset = 0
        if totals:
            dashboard.write_row(row_offset, 0, ["Column", "Total"])
            row_offset += 1
            for column in totals:
                if column in df.columns:
                    if pd.api.types.is_numeric_dtype(df[column]):
                        total_value = df[column].sum()
                    else:
                        total_value = df[column].value_counts().sum()
                    dashboard.write_row(row_offset, 0, [column, total_value])
                    row_offset += 1
            row_offset += 1

        for chart_type, config in chart_config.items():
            category_col = config.get('category_col')
            value_col = config.get('value_col')
            if not category_col or not value_col:
                continue
            if pd.api.types.is_numeric_dtype(df[value_col]):
                summary = df.groupby(category_col)[value_col].sum().reset_index()
            else:
                summary = df[category_col].value_counts().reset_index()
                summary.columns = [category_col, value_col]
            dashboard.write_row(row_offset, 0, [category_col, value_col])  # Write header
            for idx, row in enumerate(summary.itertuples(index=False), start=1):
                dashboard.write_row(row_offset + idx, 0, row)

            chart = None
            if chart_type == "bar":
                chart = workbook.add_chart({'type': 'column'})
            elif chart_type == "line":
                chart = workbook.add_chart({'type': 'line'})
            elif chart_type == "pie":
                chart = workbook.add_chart({'type': 'pie'})
                chart.set_style(10)
            elif chart_type == "doughnut":
                chart = workbook.add_chart({'type': 'doughnut'})
                chart.set_style(10)

            chart.add_series({
                'name': f'{value_col} by {category_col}',
                'categories': [f'Dashboard', row_offset + 1, 0, row_offset + len(summary), 0],
                'values': [f'Dashboard', row_offset + 1, 1, row_offset + len(summary), 1],
                'data_labels': {'value': True, 'category': True},
            })

            if chart_type in ["bar", "line"]:
                chart.set_x_axis({'name': category_col, 'name_font': {'size': 12, 'bold': True}})
                chart.set_y_axis({'name': value_col, 'name_font': {'size': 12, 'bold': True}})
            elif chart_type in ["pie", "doughnut"]:
                chart.set_title({'name': f'{value_col} by {category_col}'})

            if chart:
                dashboard.insert_chart(row_offset, 3, chart, {'x_scale': 1.5, 'y_scale': 1.5})
            row_offset += len(summary) + 5

    print(f"Excel file with dashboard saved at: {file_path}")

def fetch_kobo_data(token, form_id, base_url="https://kf.kobotoolbox.org/api/v2"):
    """
    Fetch data from KoBoToolbox for a specified form and load it into a DataFrame using KoboExtractor.
    
    Parameters:
    - token (str): API token for KoBoToolbox.
    - form_id (str): The unique identifier of the form to fetch data from.
    - base_url (str): The base URL for the KoBoToolbox API. Default is for KoBoToolbox.

    Returns:
    - df (pandas.DataFrame): Data from KoBoToolbox in a DataFrame format.
    """
    try:
        # Initialize KoboExtractor with token and base URL
        kobo = KoboExtractor(token, base_url)
        
        # Fetch the data for the specified form
        print("Fetching data from KoBoToolbox...")
        data = kobo.get_data(form_id)
        
        # Convert the data to a DataFrame
        df = pd.json_normalize(data['results'])
        
        print("Data fetched successfully!")
        return df

    except Exception as err:
        print(f"Error fetching data: {err}")

def fetch_surveycto_data(isDataset, servername, form_or_dataset_id, username, password):
    """
    Fetch data from SurveyCTO for a specified form or dataset and load it into a DataFrame.
    
    Parameters:
    - isDataset (bool): If True, fetches data from a dataset; if False, fetches data from a form.
    - servername (str): The SurveyCTO server name (without "https://").
    - form_or_dataset_id (str): The unique ID of the form or dataset to fetch data from.
    - username (str): The SurveyCTO username for authentication.
    - password (str): The SurveyCTO password for authentication.

    Returns:
    - df (pandas.DataFrame): Data from SurveyCTO in a DataFrame format.
    """
    if isDataset:
        endpoint = f"https://{servername}.surveycto.com/api/v2/datasets/data/csv/{form_or_dataset_id}"
    else:
        endpoint = f"https://{servername}.surveycto.com/api/v1/forms/data/csv/{form_or_dataset_id}"
    
    try:
        auth = (username, password)
        
        print("Fetching data from SurveyCTO...")
        response = requests.get(endpoint, auth=auth)
        response.raise_for_status()

        df = pd.read_csv(StringIO(response.text))
        
        print("Data fetched successfully!")
        return df

    except requests.exceptions.HTTPError as http_err:
        print(f"HTTP error occurred: {http_err}")
    except Exception as err:
        print(f"Other error occurred: {err}")

def generate_session_id(df, donor_name, location_settlement, name_enumerator, submission_date, session_date, project_name):
    """
    Generates a unique session ID.

    Parameters:
        df (dataframe): DataFrame of the data
        donor_name (str): Name of the column for donor name
        location_settlement (str): Name of the column for settlement
        name_enumerator (str): Name of the column for enumerator
        submission_date (str): Name of the column for submission date
        session_date (str): Name of the column for session date
        project_name (str): Name of the project

    Returns:
        df: your initial dataframe with a new column named 'session_id_sql' where will be unique session id
    """
    required_columns = [donor_name, location_settlement, name_enumerator, submission_date, session_date]
    if not all(col in df.columns for col in required_columns):
        raise ValueError(f"The DataFrame must contain the following columns: {required_columns}")

    df['session_id_sql'] = (
        df[donor_name].str.replace(r"[ :,]", "", regex=True).str.upper().str.strip() + '-' +
        project_name + '-' +
        df[location_settlement].str.replace(r"[ :,]", "", regex=True).str.upper().str.strip() + '-' +
        df[name_enumerator].str[:3].str.replace(r"[ :,]", "", regex=True).str.upper() + '-' +
        df[submission_date].str.replace(r"[ :,]", "", regex=True).str.upper().str.strip() + '-' +
        df[session_date].str.replace(r"[ :,]", "", regex=True).str.upper().str.strip()
    )

    return df

def generate_bnf_id(name, surname, dob):
    """
    Generates a unique beneficiary ID with a hash as the final component.

    Parameters:
        name (str): First name of the person.
        surname (str): Last name of the person.
        dob (str): Date of birth in 'YYYY-MM-DD' format.

    Returns:
        str: Generated unique beneficiary ID with hash included.
    """
    # Handle missing surname or dob
    surname = surname or "UNKNOWN"
    dob = dob or "0000-00-00"

    surname_length = len(surname)
    surname_part = surname[:3].upper().ljust(3, 'X')  # Pads with 'X' if fewer than 3 letters
    name_part = name[:3].upper().ljust(3, 'X')

    # Convert DOB from 'YYYY-MM-DD' to 'DDMMYY' format if valid
    if dob != "0000-00-00":
        dob_parts = dob.split("-")
        dob_formatted = dob_parts[2] + dob_parts[1] + dob_parts[0][2:]  # DDMMYY format
    else:
        dob_formatted = "000000"  # Default DOB format

    to_hash = f'{surname}{name}{dob}'
    
    base_id = f"{surname_length}-{surname_part}-{name_part}-{dob_formatted}"
    hash_suffix = hashlib.md5(to_hash.encode()).hexdigest()
    beneficiary_id = f"{base_id}-{hash_suffix}"

    return beneficiary_id

def gen_encryption_key(password):
    """
    Generates an AES encryption key using a password and a random salt.

    Parameters:
        password (str): The password or passphrase used for key derivation.

    Returns:
        str: A formatted string showing the derived key and salt.
    """
    salt = get_random_bytes(32)
    key = PBKDF2(password, salt, dkLen=32, count=1000000)
    formatted = f"Key: {key}\nSalt: {salt}"
    return formatted

def encrypt_value(value, key):
    """
    Encrypts a given value (string or number) using AES encryption in CBC mode with a random IV.
    """
    value = str(value).encode()
    iv = os.urandom(16)
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    padder = padding.PKCS7(algorithms.AES.block_size).padder()
    padded_value = padder.update(value) + padder.finalize()
    ciphertext = encryptor.update(padded_value) + encryptor.finalize()
    encrypted_value = base64.b64encode(iv + ciphertext).decode('utf-8')
    
    return encrypted_value

def decrypt_value(encrypted_data, key):
    """
    Decrypts a given encrypted value using AES encryption in CBC mode.
    """
    encrypted_data_bytes = base64.b64decode(encrypted_data)
    iv = encrypted_data_bytes[:16]
    ciphertext = encrypted_data_bytes[16:]
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    decryptor = cipher.decryptor()
    decrypted_padded_value = decryptor.update(ciphertext) + decryptor.finalize()
    unpadder = padding.PKCS7(algorithms.AES.block_size).unpadder()
    decrypted_value = unpadder.update(decrypted_padded_value) + unpadder.finalize()
    
    return decrypted_value.decode('utf-8')

def rederive_key(password, salt):
    """
    Re-derives the AES encryption key using the original password and salt.

    Parameters:
        password (str): The original password or passphrase used for key derivation.
        salt (bytes): The original salt used during the initial key derivation.

    Returns:
        bytes: The re-derived 32-byte encryption key.
    """
    key = PBKDF2(password, salt, dkLen=32, count=1000000)
    return key

def encrypt_file(input_file_path, output_file_path, key):
    """
    Encrypts the contents of a specified file using AES encryption in CBC mode with PKCS7 padding.

    Parameters:
        input_file_path (str): The path to the file that needs to be encrypted.
        output_file_path (str): The path where the encrypted file will be saved. (.en extension)
        key (bytes): The 32-byte AES encryption key.

    Returns:
        None: This function does not return a value but saves the encrypted file at the specified path.

    Notes:
        - The function generates a random 16-byte initialization vector (IV) for each encryption operation.
        - The IV is written at the beginning of the output file and is required for decryption.
        - The file is read and encrypted in chunks to optimize memory usage.
    """
    iv = os.urandom(16)
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
    encryptor = cipher.encryptor()
    padder = padding.PKCS7(algorithms.AES.block_size).padder()
    with open(input_file_path, 'rb') as input_file, open(output_file_path, 'wb') as output_file:
        output_file.write(iv)
        while True:
            chunk = input_file.read(1024)
            if len(chunk) == 0:
                break
            padded_chunk = padder.update(chunk)
            encrypted_chunk = encryptor.update(padded_chunk)
            output_file.write(encrypted_chunk)
        padded_chunk = padder.finalize()
        encrypted_chunk = encryptor.update(padded_chunk) + encryptor.finalize()
        output_file.write(encrypted_chunk)
    print(f"File '{input_file_path}' encrypted successfully and saved as '{output_file_path}'")

def decrypt_file(encrypted_file_path, output_file_path, key):
    """
    Decrypts a file that was encrypted using AES encryption in CBC mode with PKCS7 padding.

    Parameters:
        encrypted_file_path (str): The path to the encrypted file that needs to be decrypted.
        output_file_path (str): The path where the decrypted file will be saved. (.en extension)
        key (bytes): The 32-byte AES decryption key.

    Returns:
        None: This function does not return a value but saves the decrypted file at the specified path.

    Notes:
        - The function reads the 16-byte initialization vector (IV) from the beginning of the encrypted file.
        - If the IV is missing or invalid, a ValueError will be raised.
        - After decryption, PKCS7 padding is removed to restore the original content.
        - The file is read and decrypted in chunks for efficient memory usage.
    """
    with open(encrypted_file_path, 'rb') as encrypted_file, open(output_file_path, 'wb') as output_file:
        iv = encrypted_file.read(16)
        if len(iv) != 16:
            raise ValueError("Invalid IV length, file may not be encrypted correctly.")
        cipher = Cipher(algorithms.AES(key), modes.CBC(iv), backend=default_backend())
        decryptor = cipher.decryptor()
        unpadder = padding.PKCS7(algorithms.AES.block_size).unpadder()
        while True:
            chunk = encrypted_file.read(1024)
            if len(chunk) == 0:
                break
            decrypted_chunk = decryptor.update(chunk)
            if len(chunk) < 1024:
                decrypted_chunk = unpadder.update(decrypted_chunk) + unpadder.finalize()
            output_file.write(decrypted_chunk)
        output_file.write(decryptor.finalize())
    print(f"File '{encrypted_file_path}' decrypted successfully and saved as '{output_file_path}'")

def hash_password(password):
    """
    Hashes the provided password using bcrypt and returns the resulting hash as a string.

    This function generates a secure hash for the input password by:
    - Generating a salt with a cost factor of 15 rounds, enhancing the security level of the hash.
    - Hashing the password in combination with the generated salt to ensure unique hashes for identical passwords.

    Parameters:
    password (str): The plain text password to be hashed.

    Returns:
    str: The bcrypt hash of the password, encoded as a string to facilitate storage or comparison.
    """
    salt = bcrypt.gensalt(rounds=15)
    hashed = bcrypt.hashpw(password.encode(), salt)
    return hashed.decode()

def download_surveycto_photo(url, username, password):
    try:
        response = requests.get(url, auth=(username, password), stream=True)
        response.raise_for_status()
        return response.content
    except requests.exceptions.RequestException as e:
        print(f"Error downloading the photo: {e}")
        return None

def save_photo_from_bytes(photo_bytes, save_path):
    try:
        with open(save_path, 'wb') as file:
            file.write(photo_bytes)
        print(f"Photo successfully saved: {save_path}")
    except Exception as e:
        print(f"Error saving photo: {e}")
